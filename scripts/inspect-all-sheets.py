"""
Check all sheets in US and China Excel files.
"""

from openpyxl import load_workbook

files_to_check = [
    "/home/curryman/Websites/golden-pages/Country Contacts/US Government Officials.xlsx",
    "/home/curryman/Websites/golden-pages/Country Contacts/China Parliament Contact Directory.xlsx",
]

for file_path in files_to_check:
    print(f"\n{'='*80}")
    print(f"File: {file_path.split('/')[-1]}")
    print('='*80)

    wb = load_workbook(filename=file_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name} ({sheet.max_row} rows)")

        # Get headers
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value if cell.value else "")

        print(f"  Columns: {', '.join([h for h in headers if h])}")

        # Show first 2 data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=3, values_only=True), start=2):
            if row and any(row):
                row_str = ' | '.join([str(v)[:30] if v else '' for v in row])
                print(f"  Row {row_idx}: {row_str}")

    wb.close()
