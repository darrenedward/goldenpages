"""
Quick check of US Government Officials Excel file to see what columns exist.
"""

from openpyxl import load_workbook

file_path = "/home/curryman/Websites/golden-pages/Country Contacts/US Government Officials.xlsx"

wb = load_workbook(filename=file_path, read_only=True, data_only=True)
sheet = wb.active

print(f"File: US Government Officials.xlsx")
print(f"Sheet: {sheet.title}")
print(f"Rows: {sheet.max_row}, Columns: {sheet.max_column}\n")

# Get headers
headers = []
for cell in sheet[1]:
    headers.append(cell.value if cell.value else "")

print("Column Headers:")
for i, h in enumerate(headers):
    print(f"  {i+1}. {h}")

print("\nFirst 10 rows of data:")
print("-" * 80)

for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=11, values_only=True), start=2):
    # Create a dict for easier reading
    row_data = {}
    for i, val in enumerate(row):
        if i < len(headers) and headers[i]:
            row_data[headers[i]] = val

    # Print name, department, and key fields
    print(f"Row {row_idx}:")
    for key in ['Name', 'Official Name', 'Full Name', 'Department', 'Ministry', 'Organization', 'Role', 'Position', 'Title']:
        if key in row_data and row_data[key]:
            print(f"  {key}: {row_data[key]}")
    print()

wb.close()
