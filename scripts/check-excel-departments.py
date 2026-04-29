"""
Check Excel spreadsheets for department information that may have been lost during import.
"""

import os
import sys
from openpyxl import load_workbook

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def get_column_headers(sheet):
    """Get the first row as headers."""
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value if cell.value else "")
    return headers

def find_contact_in_excel(file_path, contact_name):
    """Search for a contact in an Excel file and return their row data."""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = wb.active
        headers = get_column_headers(sheet)

        # Find columns that might contain name, department, organization
        name_col = None
        dept_col = None
        org_col = None

        for i, h in enumerate(headers):
            h_lower = str(h).lower() if h else ""
            if any(x in h_lower for x in ['name', 'official', 'person', 'contact']):
                name_col = i
            if any(x in h_lower for x in ['department', 'dept', 'ministry', 'division']):
                dept_col = i
            if any(x in h_lower for x in ['organization', 'organisation', 'org', 'agency']):
                org_col = i

        # Search for the contact
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= max([c for c in [name_col, dept_col, org_col] if c is not None]):
                continue

            name_value = str(row[name_col]) if name_col is not None and row[name_col] else ""

            # Check if this matches our contact
            if contact_name.lower() in name_value.lower() or name_value.lower() in contact_name.lower():
                return {
                    'file': os.path.basename(file_path),
                    'row': row_idx,
                    'name': row[name_col] if name_col is not None else None,
                    'department': row[dept_col] if dept_col is not None else None,
                    'organization': row[org_col] if org_col is not None else None,
                    'all_data': dict(zip(headers, row)) if headers else {}
                }

        wb.close()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

    return None

# Contacts with NULL departments from the database
contacts_to_check = [
    'Mitt Romney',
    'Joe Manchin',
    'Li Qiang',
    'Xi Jinping',
]

base_dir = "/home/curryman/Websites/golden-pages/Country Contacts"

print("Searching Excel files for department information...\n")

found_results = []

for contact_name in contacts_to_check:
    print(f"Searching for: {contact_name}")
    found = False

    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if file.endswith(('.xlsx', '.xls')):
                file_path = os.path.join(root, file)
                result = find_contact_in_excel(file_path, contact_name)

                if result:
                    found = True
                    found_results.append(result)
                    print(f"  ✅ Found in: {result['file']} (Row {result['row']})")
                    print(f"     Name: {result['name']}")
                    print(f"     Department: {result['department']}")
                    print(f"     Organization: {result['organization']}")
                    print()

    if not found:
        print(f"  ❌ Not found in any Excel file\n")

print("\n" + "="*60)
print("SUMMARY")
print("="*60)
print(f"Found {len(found_results)} contacts with department info in Excel files")
